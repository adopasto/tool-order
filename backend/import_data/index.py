"""Import podkladov do evidencie - mirror src/import/index.js.

    python -m import_data.index              # spracuje vsetko v priecinku import/
    python -m import_data.index subor.xlsx    # spracuje jeden subor

Import je idempotentny: polozky sa paruju podla kodu, ponuky podla cisla
a dodavatela. Opakovane spustenie preto nic nezduplikuje, len doplni zmeny.
Skladove mnozstva import nikdy nemeni - tie vznikaju vyhradne pohybmi.
"""
from __future__ import annotations

import json
import re
import shutil
import sys
import unicodedata
from pathlib import Path
from typing import Optional

import openpyxl

from app import config, db

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")  # konzola na Windows (cp1250) by inak padla na diakritike

IMPORT_DIR = Path(__file__).resolve().parent.parent.parent / "import"
QUOTES_DIR = config.QUOTES_DIR

_log: list[str] = []


def zapis(t: str) -> None:
    _log.append(t)
    print(t)


# ---------------------------------------------------------------- pomocne

def nn(v) -> Optional[str]:
    if v is None or str(v).strip() == "":
        return None
    return str(v).strip()


def cislo(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def kod_z_nazvu(nazov: str) -> str:
    """Kod z nazvu pre polozky, ktore v prehlade kod nemaju."""
    bez_diakritiky = "".join(
        c for c in unicodedata.normalize("NFD", nazov) if unicodedata.category(c) != "Mn"
    )
    kod = re.sub(r"[^A-Z0-9]+", "-", bez_diakritiky.upper()).strip("-")
    return kod[:20]


def kategoria_podla_nazvu(nazov: str) -> Optional[str]:
    """Zaradenie do kategorie podla nazvu - prehlad kategoriu neobsahuje."""
    n = nazov.lower()
    if re.search(r"hrot|stpc|sttc|xds|xnt|xht|\brt\b|\bxt\b", n):
        return "Spájkovacie hroty"
    if re.search(r"\bpero\b|odsavacie pero|wxp|wxmp|wxdp", n):
        return "Spájkovacia technika"
    if re.search(r"hubka|špóny|spony|filter|activator|pásik|pasik|cybersolv|kefa|štetec|stetec", n):
        return "Spotrebný materiál"
    if re.search(r"kliešte|kliste|nož|noz|skrutkovač|skrutkovac|pinzeta|erem|ptr-|pnr-|dp-", n):
        return "Ručné náradie"
    if re.search(r"rukavic|okuliare|plášť|plast", n):
        return "OOPP"
    return None


def kategoria_id(nazov: Optional[str]) -> Optional[int]:
    if not nazov:
        return None
    k = db.get("SELECT id FROM categories WHERE lower(name) = lower(?)", (nazov,))
    if k:
        return k["id"]
    poradie = db.get("SELECT COALESCE(MAX(sort_order),0)+10 AS s FROM categories")["s"]
    cat_id = db.run("INSERT INTO categories (name, sort_order) VALUES (?,?)", (nazov, poradie))["lastrowid"]
    zapis(f"  + kategória: {nazov}")
    return cat_id


def uloz_polozku(p: dict) -> dict:
    """Zalozi alebo doplni polozku podla kodu. Nikdy nemeni stock_qty."""
    existuje = db.get("SELECT * FROM items WHERE upper(code) = upper(?)", (p["code"],))

    if not existuje:
        item_id = db.run(
            """INSERT INTO items
               (code, name, description, category_id, unit, reorder_point, reorder_qty,
                is_esd, active, usage_6m, ref_price, ref_price_note)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (p["code"], p["name"], p.get("description"), p.get("category_id"), p.get("unit") or "ks",
             p.get("reorder_point", 5), p.get("reorder_qty"), 1 if p.get("is_esd") else 0,
             0 if p.get("active") == 0 else 1, p.get("usage_6m"), p.get("ref_price"),
             p.get("ref_price_note")),
        )["lastrowid"]
        return {"id": item_id, "novy": True}

    # Existujucu polozku len doplname - rucne upravy v aplikacii maju prednost.
    db.run(
        """UPDATE items SET
             name = COALESCE(?, name),
             description = COALESCE(description, ?),
             category_id = COALESCE(category_id, ?),
             usage_6m = COALESCE(?, usage_6m),
             ref_price = COALESCE(?, ref_price),
             ref_price_note = COALESCE(?, ref_price_note)
           WHERE id = ?""",
        (p.get("name"), p.get("description"), p.get("category_id"),
         p.get("usage_6m"), p.get("ref_price"), p.get("ref_price_note"), existuje["id"]),
    )
    return {"id": existuje["id"], "novy": False}


# ------------------------------------------------- prehlad zasob z tabulky

def import_prehlad(subor: Path) -> None:
    """Ocakavane stlpce (poradie nerozhoduje, hlada sa podla nazvu):
       Kod | Nazov | Spotreba 6 mes. | Min. zasoba | Cena / ks | Navrh 6M zasoba | Poznamka"""
    wb = openpyxl.load_workbook(subor, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        zapis("  hárok je prázdny")
        return

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    riadky = [dict(zip(header, row)) for row in rows[1:] if any(v is not None for v in row)]
    if not riadky:
        zapis("  hárok je prázdny")
        return

    def najdi(vzor: str) -> Optional[str]:
        rx = re.compile(vzor)
        for k in header:
            if rx.search(re.sub(r"\s+", " ", k).strip().lower()):
                return k
        return None

    K = {
        "kod": najdi(r"^kód"), "nazov": najdi(r"^názov"), "spotreba": najdi(r"spotreba"),
        "min": najdi(r"min\.? zásoba"), "cena": najdi(r"cena"), "navrh": najdi(r"návrh"),
        "pozn": najdi(r"poznámka"),
    }
    if not K["nazov"]:
        zapis("  ! nenašiel som stĺpec Názov, súbor preskakujem")
        return

    zdroj = subor.name
    nove = 0
    doplnene = 0
    bez_kodu: list[str] = []

    with db.transaction():
        for r in riadky:
            nazov = nn(r.get(K["nazov"]))
            if not nazov:
                continue

            kod = nn(r.get(K["kod"])) if K["kod"] else None
            vlastny_kod = not kod
            if not kod:
                kod = kod_z_nazvu(nazov)
                bez_kodu.append(f"{kod} ({nazov})")

            pozn = nn(r.get(K["pozn"])) if K["pozn"] else None
            balenie = bool(pozn and re.search(r"balík|balik|balen", pozn, re.I))

            cena = cislo(r.get(K["cena"])) if K["cena"] else None
            reorder_point = (cislo(r.get(K["min"])) if K["min"] else None)
            if reorder_point is None:
                reorder_point = 5
            v = uloz_polozku({
                "code": kod,
                "name": nazov,
                "description": pozn,
                "category_id": kategoria_id(kategoria_podla_nazvu(nazov)),
                "unit": "bal" if balenie else "ks",
                "reorder_point": reorder_point,
                "reorder_qty": cislo(r.get(K["navrh"])) if K["navrh"] else None,
                "is_esd": 1 if re.search(r"esd", nazov, re.I) else 0,
                # polozky bez kodu su neuplne - do katalogu ich nepustame, kym ich niekto neskontroluje
                "active": 0 if vlastny_kod else 1,
                "usage_6m": cislo(r.get(K["spotreba"])) if K["spotreba"] else None,
                "ref_price": cena,
                "ref_price_note": f"orientačne podľa {zdroj}" if cena is not None else None,
            })
            if v["novy"]:
                nove += 1
            else:
                doplnene += 1

    zapis(f"  položky: {nove} nových, {doplnene} doplnených")
    if bez_kodu:
        zapis(f"  ! {len(bez_kodu)} položiek nemalo v prehľade kód – založené ako NEAKTÍVNE")
        zapis("    doplň im kód a cenu v Správe → Položky, potom ich aktivuj:")
        for b in bez_kodu:
            zapis(f"      {b}")


# ------------------------------------------------------- cenova ponuka JSON

def uloz_dodavatela(d: dict) -> int:
    existuje = db.get(
        "SELECT * FROM suppliers WHERE lower(name) = lower(?) OR (ico IS NOT NULL AND ico = ?)",
        (d["nazov"], d.get("ico")),
    )
    if existuje:
        db.run(
            """UPDATE suppliers SET ico = COALESCE(ico, ?), dic = COALESCE(dic, ?),
                 contact_person = COALESCE(contact_person, ?), email = COALESCE(email, ?),
                 phone = COALESCE(phone, ?), address = COALESCE(address, ?),
                 note = COALESCE(note, ?) WHERE id = ?""",
            (d.get("ico"), d.get("dic"), d.get("kontakt"), d.get("email"),
             d.get("telefon"), d.get("adresa"), d.get("poznamka"), existuje["id"]),
        )
        return existuje["id"]

    supplier_id = db.run(
        """INSERT INTO suppliers
           (name, ico, dic, contact_person, email, phone, address, lead_time_days, note, active)
           VALUES (?,?,?,?,?,?,?,?,?,1)""",
        (d["nazov"], d.get("ico"), d.get("dic"), d.get("kontakt"), d.get("email"),
         d.get("telefon"), d.get("adresa"), d.get("dodacia_lehota", 7), d.get("poznamka")),
    )["lastrowid"]
    zapis(f"  + dodávateľ: {d['nazov']}")
    return supplier_id


def import_ponuka(subor: Path) -> None:
    p = json.loads(subor.read_text(encoding="utf-8"))
    supplier_id = uloz_dodavatela(p["dodavatel"])

    # priloha (PDF) sa odklada k databaze, aby ju aplikacia vedela ponuknut na stiahnutie
    priloha = None
    if p.get("priloha"):
        zdroj = IMPORT_DIR / "prilohy" / p["priloha"]
        if zdroj.exists():
            QUOTES_DIR.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(zdroj, QUOTES_DIR / p["priloha"])
            priloha = p["priloha"]
        else:
            zapis(f"  ! príloha {p['priloha']} sa nenašla v import/prilohy")

    with db.transaction():
        stara = db.get("SELECT id FROM quotes WHERE supplier_id = ? AND number = ?", (supplier_id, p["cislo"]))
        if stara:
            db.run("DELETE FROM quotes WHERE id = ?", (stara["id"],))  # prepiseme celu

        quote_id = db.run(
            """INSERT INTO quotes
               (number, supplier_id, issued_at, valid_until, total_no_vat, total_vat,
                currency, file_path, note)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (p["cislo"], supplier_id, p.get("vystavena"), p.get("platna_do"),
             p.get("spolu_bez_dph"), p.get("spolu_s_dph"), p.get("mena") or "EUR",
             priloha, p.get("poznamka")),
        )["lastrowid"]

        for r in p["riadky"]:
            item_id = None
            polozka = r.get("polozka")
            if polozka:
                v = uloz_polozku({
                    "code": polozka["kod"],
                    "name": polozka["nazov"],
                    "description": polozka.get("popis"),
                    "category_id": kategoria_id(polozka.get("kategoria")),
                    "unit": polozka.get("jednotka") or "ks",
                    "reorder_point": polozka.get("signalna_zasoba", 5),
                    "is_esd": 1 if polozka.get("esd") else 0,
                })
                item_id = v["id"]
                if v["novy"]:
                    zapis(f"  + položka: {polozka['kod']} – {polozka['nazov']}")

                # vazba na dodavatela: cena najnizsieho pasma = bezna objednavacia cena
                zaklad = min(r["pasma"], key=lambda x: x["od_mnozstva"]) if r["pasma"] else None
                db.run(
                    """INSERT INTO supplier_items
                       (supplier_id, item_id, supplier_sku, price, min_order_qty, is_primary)
                       VALUES (?,?,?,?,?,1)
                       ON CONFLICT(supplier_id, item_id) DO UPDATE SET
                         supplier_sku = excluded.supplier_sku, price = excluded.price,
                         min_order_qty = excluded.min_order_qty""",
                    (supplier_id, item_id, r.get("kod_dodavatela"),
                     zaklad["cena"] if zaklad else None, zaklad["od_mnozstva"] if zaklad else 1),
                )

            for pasmo in r["pasma"]:
                db.run(
                    """INSERT INTO quote_lines
                       (quote_id, item_id, supplier_sku, description, min_qty, unit, unit_price, note)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (quote_id, item_id, r.get("kod_dodavatela"), r.get("popis"),
                     pasmo["od_mnozstva"], (polozka or {}).get("jednotka"), pasmo["cena"],
                     pasmo.get("poznamka")),
                )

    pasiem = sum(len(r["pasma"]) for r in p["riadky"])
    priloha_txt = f", príloha {priloha}" if priloha else ""
    zapis(f"  ponuka {p['cislo']}: {len(p['riadky'])} položiek, {pasiem} cenových pásiem{priloha_txt}")


# ------------------------------------------------------------------- beh

def spusti(subory: list[Path]) -> None:
    db.migrate()
    for f in subory:
        zapis(f"\n▸ {f.name}")
        try:
            if re.search(r"\.xlsx?$", f.name, re.I):
                import_prehlad(f)
            elif re.search(r"\.json$", f.name, re.I):
                import_ponuka(f)
            else:
                zapis("  neznámy typ súboru, preskakujem")
        except Exception as e:
            zapis(f"  ! chyba: {e}")

    s = db.get("""SELECT
        (SELECT COUNT(*) FROM items) AS polozky,
        (SELECT COUNT(*) FROM items WHERE active = 0) AS neaktivne,
        (SELECT COUNT(*) FROM suppliers) AS dodavatelia,
        (SELECT COUNT(*) FROM quotes) AS ponuky,
        (SELECT COUNT(*) FROM quote_lines) AS pasma""")
    zapis(f"\nStav evidencie: {s['polozky']} položiek (z toho {s['neaktivne']} neaktívnych), "
          f"{s['dodavatelia']} dodávateľov, {s['ponuky']} ponúk, {s['pasma']} cenových pásiem.")


def main() -> None:
    args = sys.argv[1:]
    if args:
        subory = [Path(a).resolve() for a in args]
    elif IMPORT_DIR.exists():
        subory = sorted(f for f in IMPORT_DIR.iterdir() if re.search(r"\.(xlsx?|json)$", f.name, re.I))
    else:
        subory = []

    if not subory:
        print(f"Nič na import. Vlož súbory do {IMPORT_DIR} (xlsx prehľady, json ponuky).")
        return
    spusti(subory)


if __name__ == "__main__":
    main()
